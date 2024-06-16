import re

import pandas as pd
from utils.data_operations import rename_columns_to_snake_case
from openpyxl import load_workbook


def to_snake_case(name):
    """Convert CamelCase to snake_case."""
    s1 = re.sub('(.)([A-Z][a-z]+)', r'\1_\2', name)
    return re.sub('([a-z0-9])([A-Z])', r'\1_\2', s1).lower()


def read_excel_files(file_path, skiprows, has_headers):
    xls = pd.ExcelFile(file_path)
    all_dfs = []

    for sheet_name in xls.sheet_names:
        print(f"Reading sheet: {sheet_name} from file: {file_path}")
        df = pd.read_excel(xls, sheet_name=sheet_name, skiprows=skiprows, usecols="A:AK")

        empty_row_index = df.isnull().all(axis=1).idxmax()
        df = df.iloc[:empty_row_index] if not pd.isnull(empty_row_index) else df

        df = fill_merged_rows(file_path, sheet_name, df)

        if not has_headers:
            yield df
        else:
            yield df[1:]


def fill_merged_rows(file_path, sheet_name, df):
    wb = load_workbook(file_path, data_only=True)
    sheet = wb.active
    all_data = []

    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
        row_data = []
        for cell in row:
            if cell.coordinate in sheet.merged_cells:
                # Skip the merged cell, as its value has already been processed
                continue
            if cell.value:
                row_data.append(cell.value)
            else:
                row_data.append(None)
        all_data.append(row_data)

    return all_data


